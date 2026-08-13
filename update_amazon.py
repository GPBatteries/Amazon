import os
import io
import json
import time
import datetime as dt

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import crypto_utils

# ----------------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------------
SHEET_ID = "19h1OUyvNQbIVxNIkPEMUsHM8pLNPuw0jMHKkUQiB1l8"
DAILY_IMP_GID = "210013716"
PRODUCTS_FILE = "products.json"

# Aannames (bewerkbaar in de Excel; dit zijn de startwaarden) -- gelden nu
# nog voor alle producten gelijk. Per-product aannames kunnen later aan
# products.json toegevoegd worden als daar behoefte aan is.
VAT_RATE = 0.20                      # net = gross / (1 + BTW) -- geldt voor alle producten gelijk
COMMISSION_ADFEE_PCT_DEFAULT = 0.153 # 15,3% van gross RSP -- fallback als een product geen eigen waarde heeft
FBA_DEFAULT = 2.42                   # GBP per stuk -- fallback
COGS_DEFAULT = 3.96                  # GBP per stuk -- fallback

CSV_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={DAILY_IMP_GID}"

NEEDED = ["date", "childAsin", "unitsOrdered", "orderedProductSales", "unitSessionPercentage"]

FONT = "Arial"
BLUE = "0000FF"
GREY = "808080"


def load_products() -> list[dict]:
    with open(PRODUCTS_FILE, encoding="utf-8") as fh:
        data = json.load(fh)
    products = data.get("products", [])
    if not products:
        raise SystemExit(f"Geen producten gevonden in {PRODUCTS_FILE}.")
    return products


def resolve_costs(product: dict, date_str: str) -> tuple[float, float, float]:
    """
    Bepaalt welke fba/cogs/commissionPct gelden voor een specifieke datum,
    in deze volgorde van specificiteit:
      1. costHistory-item met de meest recente effectiveFrom <= date_str
         (dus: "de laatste wijziging die op deze datum al was ingegaan").
      2. De vlakke fba/cogs/commissionPct op het product zelf (geen datum
         gekoppeld -- de "gewone" instelling via het snelle formulier).
      3. De globale fallback-defaults (FBA_DEFAULT/COGS_DEFAULT/COMMISSION_ADFEE_PCT_DEFAULT).

    Elk veld (fba/cogs/commissionPct) wordt AFZONDERLIJK opgelost -- een
    costHistory-item dat bv. alleen fba wijzigt, laat cogs/commissionPct via
    de lagere niveaus oplossen (dus je hoeft niet steeds alle 3 mee te geven).
    """
    history = sorted(product.get("costHistory", []), key=lambda h: h["effectiveFrom"])
    applicable = [h for h in history if h["effectiveFrom"] <= date_str]
    latest = applicable[-1] if applicable else {}

    fba = latest.get("fba", product.get("fba", FBA_DEFAULT))
    cogs = latest.get("cogs", product.get("cogs", COGS_DEFAULT))
    commission_pct = latest.get("commissionPct", product.get("commissionPct", COMMISSION_ADFEE_PCT_DEFAULT))
    return float(fba), float(cogs), float(commission_pct)

# ----------------------------------------------------------------------------
# Data inlezen
# ----------------------------------------------------------------------------
def decrypt_history_csv(enc_path: str) -> pd.DataFrame:
    """Ontsleutelt een history.csv.enc-bestand (zie crypto_utils) naar een DataFrame."""
    password = crypto_utils.get_site_password()
    with open(enc_path, "rb") as fh:
        blob = fh.read()
    plaintext = crypto_utils.decrypt_bytes(blob, password).decode("utf-8")
    return pd.read_csv(io.StringIO(plaintext))


def load_daily_imp() -> pd.DataFrame:
    spapi_history_enc = os.path.join("output", "spapi_history.csv.enc")
    local = os.environ.get("AMAZON_LOCAL_CSV")
    use_spapi = os.environ.get("DATA_SOURCE", "").lower() == "spapi"

    if use_spapi:
        if not os.path.exists(spapi_history_enc):
            raise SystemExit(
                f"DATA_SOURCE=spapi maar {spapi_history_enc} bestaat nog niet. "
                "Draai eerst fetch_spapi_daily.py."
            )
        df = decrypt_history_csv(spapi_history_enc)
    elif local:
        df = pd.read_csv(local)
    else:
        # Cache-buster + no-cache headers: dwingt Google een verse export te geven
        # in plaats van een gecachte versie die de nieuwste rij nog mist.
        url = f"{CSV_URL}&_cb={int(time.time())}"
        r = requests.get(url, timeout=60,
                         headers={"Cache-Control": "no-cache", "Pragma": "no-cache"})
        r.raise_for_status()
        df = pd.read_csv(io.StringIO(r.text))

    missing = [c for c in NEEDED if c not in df.columns]
    if missing:
        raise SystemExit(f"Ontbrekende kolommen in Daily IMP: {missing}")

    df = df[NEEDED].copy()
    df["date"] = pd.to_datetime(df["date"]).dt.date
    for c in ["unitsOrdered", "orderedProductSales", "unitSessionPercentage"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    # Daily IMP bevat per datum dubbele rijen (een blok zonder en een met traffic).
    # units/sales zijn identiek in beide; dedup naar 1 rij per (datum, ASIN) via max.
    # Voor CVR pakt max automatisch de rij met traffic (rij zonder traffic = 0).
    df = (df.groupby(["date", "childAsin"], as_index=False)
            .agg({"unitsOrdered": "max",
                  "orderedProductSales": "max",
                  "unitSessionPercentage": "max"}))
    df = df.sort_values(["childAsin", "date"]).reset_index(drop=True)
    return df

# ----------------------------------------------------------------------------
# Excel bouwen
# ----------------------------------------------------------------------------
def build(df: pd.DataFrame, asin: str, fba: float, cogs: float, commission_pct: float):
    output_path = os.path.join("output", f"Amazon_{asin}.xlsx")
    wb = Workbook()

    # --- Data-tab (ruwe bron, SUMIFS leest hieruit) ---
    data = wb.active
    data.title = "Data"
    data.append(NEEDED)
    for _, row in df.iterrows():
        data.append([row["date"], row["childAsin"],
                     float(row["unitsOrdered"]), float(row["orderedProductSales"]),
                     float(row["unitSessionPercentage"])])
    for r in range(2, data.max_row + 1):
        data.cell(r, 1).number_format = "yyyy-mm-dd"
    for c in range(1, 6):
        data.cell(1, c).font = Font(name=FONT, bold=True)
    data.sheet_state = "hidden"

    # --- Rapport-tab ---
    rep = wb.create_sheet("Rapport")
    dates = sorted(df["date"].unique())

    labels = [
        (3, "Sales #", "Column D"),
        (4, "Sales \u00a3", "Column F"),
        (5, "Avg. RSP (gross)", ""),
        (6, "Avg. RSP (net)", ""),
        (7, "Commission + AdFee", ""),
        (8, "FBA", ""),
        (9, "COGS", ""),
        (10, "Margin (%)", ""),
        (11, "Margin (abs)", ""),
        (12, "Margin Tot. (abs)", ""),
        (13, "Margin Net. (abs)", ""),
        (14, "CVR", "Column AF"),
    ]

    rep["C1"] = "Datum (Column A)"
    rep["A2"] = "ASin"
    rep["B2"] = asin
    rep["A3"] = "C40"
    for r, name, src in labels:
        rep.cell(r, 2, name)
        if src:
            rep.cell(r, 3, src)

    # Aannames (bewerkbaar)
    rep["B16"] = "Aannames"
    assum = [
        (17, "BTW", VAT_RATE, "0.0%"),
        (18, "Commission + AdFee %", commission_pct, "0.0%"),
        (19, "FBA (\u00a3/stuk)", fba, "\u00a3#,##0.00"),
        (20, "COGS (\u00a3/stuk)", cogs, "\u00a3#,##0.00"),
    ]
    for r, name, val, fmt in assum:
        rep.cell(r, 2, name)
        c = rep.cell(r, 3, val)
        c.number_format = fmt
        c.font = Font(name=FONT, color=BLUE)  # blauw = bewerkbare input

    # Datakolommen per dag
    first_col = 4  # kolom D
    for i, d in enumerate(dates):
        ci = first_col + i
        L = get_column_letter(ci)
        dcell = rep.cell(1, ci, d)
        dcell.number_format = "yyyy-mm-dd"
        dcell.font = Font(name=FONT, bold=True)
        dcell.alignment = Alignment(horizontal="center")

        f = {
            3:  f"=SUMIFS(Data!$C:$C,Data!$A:$A,{L}$1,Data!$B:$B,$B$2)",
            4:  f"=SUMIFS(Data!$D:$D,Data!$A:$A,{L}$1,Data!$B:$B,$B$2)",
            5:  f"=IF({L}3=0,0,{L}4/{L}3)",
            6:  f"={L}5/(1+$C$17)",
            7:  f"=$C$18*{L}5",
            8:  "=$C$19",
            9:  "=$C$20",
            10: f"=IF({L}6=0,0,{L}11/{L}6)",
            11: f"={L}6-{L}7-{L}8-{L}9",
            12: f"={L}11*{L}3",
            13: None,  # Margin Net. (abs) bewust leeg
            14: f"=SUMIFS(Data!$E:$E,Data!$A:$A,{L}$1,Data!$B:$B,$B$2)",
        }
        for r, formula in f.items():
            if formula is None:
                continue
            rep.cell(r, ci, formula)

    # Opmaak per rij
    money = "\u00a3#,##0.00"
    fmt_by_row = {3: "#,##0", 4: money, 5: money, 6: money, 7: money,
                  8: money, 9: money, 10: "0.0%", 11: money, 12: money,
                  13: money, 14: '0.0"%"'}
    last_col = first_col + len(dates) - 1
    for r, fmt in fmt_by_row.items():
        for ci in range(first_col, last_col + 1):
            rep.cell(r, ci).number_format = fmt
            rep.cell(r, ci).font = Font(name=FONT)

    # Labels/headers opmaak
    rep["A2"].font = Font(name=FONT, bold=True)
    rep["B2"].font = Font(name=FONT, bold=True)
    rep["A3"].font = Font(name=FONT, color=GREY)
    rep["C1"].font = Font(name=FONT, italic=True, color=GREY)
    rep["B16"].font = Font(name=FONT, bold=True)
    for r, *_ in labels:
        rep.cell(r, 2).font = Font(name=FONT, bold=(r in (3, 4)))
        rep.cell(r, 3).font = Font(name=FONT, italic=True, color=GREY)

    # Kolombreedtes
    rep.column_dimensions["A"].width = 6
    rep.column_dimensions["B"].width = 20
    rep.column_dimensions["C"].width = 14
    for ci in range(first_col, last_col + 1):
        rep.column_dimensions[get_column_letter(ci)].width = 12

    rep.freeze_panes = "D2"

    # Forceer Excel om bij openen alle formules te herberekenen (anders blijven de
    # cellen leeg omdat openpyxl geen voorberekende waarden meeschrijft).
    wb.calculation.fullCalcOnLoad = True

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return len(dates), output_path

# ----------------------------------------------------------------------------
# Cijfers berekenen (identiek aan de Excel-formules) voor data.json / dashboard
# ----------------------------------------------------------------------------
def load_ads_spend() -> dict:
    """
    Leest output/ads_spend_history.csv.enc (indien aanwezig, gevuld door
    fetch_ads_spend.py) en geeft een dict {(datum, childAsin): rij} terug.
    Ontbreekt dit bestand nog (bv. eerste keer, of de losse Ads-workflow heeft
    nog niet gedraaid), dan geeft dit gewoon een lege dict terug -- niks
    breekt, de ads-kolommen blijven dan leeg in het dashboard.
    """
    path = os.path.join("output", "ads_spend_history.csv.enc")
    if not os.path.exists(path):
        return {}
    ads_df = decrypt_history_csv(path)
    return {(str(r["date"]), r["childAsin"]): r for _, r in ads_df.iterrows()}


def compute_rows(df: pd.DataFrame, ads_spend: dict, asin: str, product: dict):
    ads_spend = ads_spend or {}
    rows = []
    for _, r in df.iterrows():
        date_str = str(r["date"])
        fba, cogs, commission_pct = resolve_costs(product, date_str)

        units = float(r["unitsOrdered"])
        sales = float(r["orderedProductSales"])
        gross = sales / units if units else 0.0
        net = gross / (1 + VAT_RATE)
        commission = commission_pct * gross
        margin_abs = net - commission - fba - cogs
        margin_pct = (margin_abs / net) if net else 0.0

        row = {
            "date": date_str,
            "asin": asin,
            "units": units,
            "sales": sales,
            "grossRsp": gross,
            "netRsp": net,
            "commission": commission,
            "fba": fba,
            "cogs": cogs,
            "marginPct": margin_pct,
            "marginAbs": margin_abs,
            "marginTot": margin_abs * units,
            "cvr": float(r["unitSessionPercentage"]),
        }

        # Ads-spend is optioneel en komt uit een losse databron (Ads API,
        # via fetch_ads_spend.py) -- alleen invullen als deze (datum, asin)
        # combinatie bekend is.
        ads_row = ads_spend.get((date_str, asin))
        if ads_row is not None:
            ad_spend = float(ads_row["adSpend"])
            ad_sales14d = float(ads_row["adSales14d"])
            row["adSpend"] = ad_spend
            row["adClicks"] = int(ads_row["adClicks"])
            row["adSales14d"] = ad_sales14d
            row["adOrders14d"] = int(ads_row["adOrders14d"])
            row["acos"] = (ad_spend / ad_sales14d) if ad_sales14d else None
            row["tacos"] = (ad_spend / sales) if sales else None

            # NET MARGIN = totale marge van die dag MINUS de advertentiekosten
            # die dag. NET MARGIN % = die net margin t.o.v. de TOTALE netto-omzet
            # van die dag (net is hier de prijs per eenheid, dus vermenigvuldigen
            # met units voor het totaal -- anders klopt de verhouding niet).
            net_total = net * units
            net_margin = row["marginTot"] - ad_spend
            row["netMargin"] = net_margin
            row["netMarginPct"] = (net_margin / net_total) if net_total else None

        rows.append(row)
    return rows


# ----------------------------------------------------------------------------
def main():
    products = load_products()
    print(f"Producten uit {PRODUCTS_FILE}: {', '.join(p['asin'] for p in products)}")

    full_df = load_daily_imp()
    if full_df.empty:
        raise SystemExit("Geen rijen gevonden in de bron (spapi_history.csv/Google Sheet leeg?).")
    ads_spend_all = load_ads_spend()
    generated = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    all_rows = []
    products_meta = []
    for p in products:
        asin, name, category = p["asin"], p.get("name", p["asin"]), p.get("category", "Overig")

        pdf = full_df[full_df["childAsin"] == asin].reset_index(drop=True)
        if pdf.empty:
            print(f"LET OP: geen data voor {asin} ({name}) -- nog niet opgehaald, of nog geen sales. "
                  f"Wordt overgeslagen in data.json totdat er data is.")
            continue

        # Voor de Excel-"Aannames"-tabel tonen we de kosten die gelden op de
        # LAATSTE (meest recente) datum -- de dag-op-dag-berekening zelf
        # (compute_rows) lost per rij de juiste, op dat moment geldende
        # kosten op, dus die klopt ook als er tussentijds een costHistory-
        # wijziging is geweest.
        last_date_str = str(max(pdf["date"]))
        fba, cogs, commission_pct = resolve_costs(p, last_date_str)

        n, output_path = build(pdf, asin, fba, cogs, commission_pct)
        rows = compute_rows(pdf, ads_spend_all, asin, p)
        all_rows.extend(rows)
        products_meta.append({
            "asin": asin,
            "name": name,
            "category": category,
            "file": os.path.basename(output_path),
            "days": n,
            "rows": int(len(pdf)),
            "last_date": str(max(pdf["date"])),
            "fba": fba,
            "cogs": cogs,
            "commissionPct": commission_pct,
        })
        print(f"OK: {output_path} ({n} dagkolommen, {len(pdf)} bronrijen voor {asin}) "
              f"[fba={fba} cogs={cogs} commissie={commission_pct*100:.1f}%]")

    if not products_meta:
        raise SystemExit("Geen van de producten in products.json heeft data -- niets gebouwd.")

    meta = {
        "generated_utc": generated,
        "products": products_meta,
        "total_rows": len(all_rows),
    }
    with open(os.path.join("output", "meta.json"), "w", encoding="utf-8") as fh:
        json.dump(meta, fh, ensure_ascii=False, indent=2)

    data = {
        "generated_utc": generated,
        "assumptions": {
            "vat": VAT_RATE,
            "commissionPctDefault": COMMISSION_ADFEE_PCT_DEFAULT,
            "fbaDefault": FBA_DEFAULT,
            "cogsDefault": COGS_DEFAULT,
        },
        "products": products_meta,
        "rows": all_rows,
    }
    with open(os.path.join("output", "data.json"), "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False)

    print(f"\nKlaar: {len(products_meta)} product(en) verwerkt, {len(all_rows)} rijen totaal in data.json.")

if __name__ == "__main__":
    main()
